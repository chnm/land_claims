import openpyxl
import sqlite3
import sys

conn = sqlite3.connect('land_claims.db')
conn.row_factory = sqlite3.Row

# Create the schema.
conn.execute('DROP TABLE IF EXISTS states')
conn.execute('''
CREATE TABLE states (
    id INTEGER PRIMARY KEY,
    state TEXT NOT NULL UNIQUE
);
''')
conn.execute('DROP TABLE IF EXISTS land_offices')
conn.execute('''
CREATE TABLE land_offices (
    id INTEGER PRIMARY KEY,
    state_id INTEGER,
    land_office TEXT NOT NULL UNIQUE,
    FOREIGN KEY (state_id) REFERENCES state (id) ON DELETE CASCADE
);
''')
conn.execute('DROP TABLE IF EXISTS claims')
conn.execute('''
CREATE TABLE IF NOT EXISTS claims (
    id INTEGER PRIMARY KEY,
    land_office_id INTEGER,
    year INTEGER NOT NULL,
    claims INTEGER NULL,
    acres REAL NULL,
    FOREIGN KEY (land_office_id) REFERENCES land_offices (id) ON DELETE CASCADE
);
''')
conn.execute('DROP TABLE IF EXISTS patents')
conn.execute('''
CREATE TABLE IF NOT EXISTS patents (
    id INTEGER PRIMARY KEY,
    land_office_id INTEGER,
    year INTEGER NOT NULL,
    patents INTEGER NULL,
    acres REAL NULL,
    FOREIGN KEY (land_office_id) REFERENCES land_offices (id) ON DELETE CASCADE
);
''')
conn.execute('DROP TABLE IF EXISTS commutations')
conn.execute('''
CREATE TABLE IF NOT EXISTS commutations (
    id INTEGER PRIMARY KEY,
    land_office_id INTEGER,
    year INTEGER NOT NULL,
    commutations INTEGER NULL,
    acres REAL NULL,
    FOREIGN KEY (land_office_id) REFERENCES land_offices (id) ON DELETE CASCADE
);
''')

wb = openpyxl.load_workbook(sys.argv[1])
valid_sheets = [str(year) for year in range(1863, 1897)]

# Process the states and land offices. Here we assume that the list of land
# offices and states on the "good keys" sheet is comprehensive.
land_office_state_map = {}
state_values = set()
land_office_values = set()
for row in wb['good keys'].iter_rows(min_row=3, max_row=211, max_col=2):
    state = row[1].value
    land_office = row[0].value
    land_office_state_map[land_office] = state
    state_values.add((state,))
conn.executemany('INSERT INTO states (state) VALUES (?)', state_values)
states = {row['state']: row['id'] for row in conn.execute('SELECT * FROM states').fetchall()}
for land_office, state in land_office_state_map.items():
    land_office_values.add((states[state], land_office))
conn.executemany('INSERT INTO land_offices (state_id, land_office) VALUES (?, ?)', land_office_values)
land_offices = {row['land_office']: row['id'] for row in conn.execute('SELECT * FROM land_offices').fetchall()}

# Process the claims, patents, and commutations.
claim_values = set()
patent_values = set()
commutation_values = set()
for ws in wb:
    if ws.title in valid_sheets:
        for row in ws.iter_rows(min_row=3, max_col=7):
            if row[0].value is not None:
                if row[0].value in land_offices:
                    land_office_id = land_offices[row[0].value]
                    year = int(ws.title)
                    claim_values.add((land_office_id, year, row[1].value, row[2].value))
                    patent_values.add((land_office_id, year, row[3].value, row[4].value))
                    commutation_values.add((land_office_id, year, row[5].value, row[6].value))
conn.executemany('INSERT INTO claims (land_office_id, year, claims, acres) VALUES (?, ?, ?, ?)', claim_values)
conn.executemany('INSERT INTO patents (land_office_id, year, patents, acres) VALUES (?, ?, ?, ?)', patent_values)
conn.executemany('INSERT INTO commutations (land_office_id, year, commutations, acres) VALUES (?, ?, ?, ?)', commutation_values)

conn.commit()
conn.close()
