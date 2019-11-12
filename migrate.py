import json
import openpyxl
import sqlite3
import sys

YEAR_FROM = 1863
YEAR_TO = 1912

conn = sqlite3.connect('mapping_the_homestead_act.db')
conn.row_factory = sqlite3.Row

# Create the schema.
conn.execute('DROP TABLE IF EXISTS states')
conn.execute('''
CREATE TABLE states (
    id INTEGER PRIMARY KEY,
    state TEXT NOT NULL UNIQUE,
    coordinates TEXT
);
''')
conn.execute('DROP TABLE IF EXISTS land_offices')
conn.execute('''
CREATE TABLE land_offices (
    id INTEGER PRIMARY KEY,
    state_id INTEGER,
    land_office TEXT NOT NULL UNIQUE,
    FOREIGN KEY (state_id) REFERENCES state (id) ON DELETE CASCADE
);
''')
conn.execute('DROP TABLE IF EXISTS stats')
conn.execute('''
CREATE TABLE IF NOT EXISTS stats (
    id INTEGER PRIMARY KEY,
    land_office_id INTEGER,
    year INTEGER NOT NULL,
    type TEXT NOT NULL,
    number INTEGER NULL,
    acres REAL NULL,
    fees REAL NULL,
    FOREIGN KEY (land_office_id) REFERENCES land_offices (id) ON DELETE CASCADE
);
''')

# Process the state polygon coordinates.
with open('us_states.json', 'r') as us_states:
    data = json.load(us_states)
state_coordinates_map = {}
for feature in data['features']:
    state = feature['properties']['name']
    coordinates = feature['geometry']['coordinates']
    state_coordinates_map[state] = json.dumps(coordinates)

# Load the workbook.
wb = openpyxl.load_workbook(sys.argv[1])
valid_sheets = [str(year) for year in range(YEAR_FROM, YEAR_TO)]

# Process the states and land offices. Here we assume that the list of land
# offices and states on the "good keys" sheet is comprehensive.
land_office_state_map = {}
state_values = set()
land_office_values = set()
for row in wb['good keys'].iter_rows(min_row=3, max_row=238, max_col=2):
    state = row[1].value
    land_office = row[0].value
    land_office_state_map[land_office] = state
    state_values.add((state, state_coordinates_map[state]))
conn.executemany('INSERT INTO states (state, coordinates) VALUES (?, ?)', state_values)
states = {row['state']: row['id'] for row in conn.execute('SELECT * FROM states').fetchall()}
for land_office, state in land_office_state_map.items():
    land_office_values.add((states[state], land_office))
conn.executemany('INSERT INTO land_offices (state_id, land_office) VALUES (?, ?)', land_office_values)
land_offices = {row['land_office']: row['id'] for row in conn.execute('SELECT * FROM land_offices').fetchall()}

# Process the claims, patents, and commutations.
claim_values = set()
patent_values = set()
commutation_sec2301_values = set()
commutation_june151880_values = set()
commutation_sec2301fees_values = set()
claim_indianland_values = set()
commutation_indianland_values = set()
patent_indianland_values = set()
for ws in wb:
    if ws.title in valid_sheets:
        for row in ws.iter_rows(min_row=3, max_col=17):
            if row[0].value is not None:
                if row[0].value in land_offices:
                    land_office_id = land_offices[row[0].value]
                    year = int(ws.title)
                    # Note that the numbers must be integers (i.e. not None or float)
                    if isinstance(row[1].value, int):
                        claim_values.add((land_office_id, year, 'claim', row[1].value, row[2].value))
                    if isinstance(row[3].value, int):
                        patent_values.add((land_office_id, year, 'patent', row[3].value, row[4].value))
                    if isinstance(row[5].value, int):
                        commutation_sec2301_values.add((land_office_id, year, 'commutation_sec2301', row[5].value, row[6].value))
                    if isinstance(row[7].value, int):
                        commutation_june151880_values.add((land_office_id, year, 'commutation_june151880', row[7].value, row[8].value))
                    if isinstance(row[9].value, (int, float)):
                        commutation_sec2301fees_values.add((land_office_id, year, 'commutation_sec2301fees', row[9].value))
                    if isinstance(row[11].value, int):
                        claim_indianland_values.add((land_office_id, year, 'claim_indianland', row[11].value, row[12].value))
                    if isinstance(row[13].value, int):
                        commutation_indianland_values.add((land_office_id, year, 'commutation_indianland', row[13].value, row[14].value))
                    if isinstance(row[15].value, int):
                        patent_indianland_values.add((land_office_id, year, 'patent_indianland', row[15].value, row[16].value))

conn.executemany('INSERT INTO stats (land_office_id, year, type, number, acres) VALUES (?, ?, ?, ?, ?)', claim_values)
conn.executemany('INSERT INTO stats (land_office_id, year, type, number, acres) VALUES (?, ?, ?, ?, ?)', patent_values)
conn.executemany('INSERT INTO stats (land_office_id, year, type, number, acres) VALUES (?, ?, ?, ?, ?)', commutation_sec2301_values)
conn.executemany('INSERT INTO stats (land_office_id, year, type, number, acres) VALUES (?, ?, ?, ?, ?)', commutation_june151880_values)
conn.executemany('INSERT INTO stats (land_office_id, year, type, fees) VALUES (?, ?, ?, ?)', commutation_sec2301fees_values)
conn.executemany('INSERT INTO stats (land_office_id, year, type, number, acres) VALUES (?, ?, ?, ?, ?)', claim_indianland_values)
conn.executemany('INSERT INTO stats (land_office_id, year, type, number, acres) VALUES (?, ?, ?, ?, ?)', commutation_indianland_values)
conn.executemany('INSERT INTO stats (land_office_id, year, type, number, acres) VALUES (?, ?, ?, ?, ?)', patent_indianland_values)

conn.commit()
conn.close()
